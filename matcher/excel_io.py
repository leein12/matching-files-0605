"""엑셀 입출력."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from matcher.candidate import MatchCandidate, build_b_index, find_candidates_for_row


def _ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def validate_columns(df: pd.DataFrame, filename: str, required_columns: list[str]) -> None:
    """필수 컬럼 존재 여부를 검증한다."""
    missing = [c for c in required_columns if c not in df.columns]
    if missing:
        raise ValueError(f"{filename}에 필수 컬럼이 없습니다: {', '.join(missing)}")


def read_input_excel(
    path: Path,
    required_columns: list[str] | None = None,
) -> pd.DataFrame:
    """입력 엑셀을 읽는다."""
    if not path.exists():
        raise FileNotFoundError(f"입력 파일이 없습니다: {path}")
    df = pd.read_excel(path, dtype=str)
    df = df.fillna("")
    if required_columns is not None:
        validate_columns(df, path.name, required_columns)
    return df


def _safe_value(value: Any) -> Any:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return value


def _build_output_row(a_row: pd.Series, candidates: list[MatchCandidate]) -> dict[str, Any]:
    """결과 행을 구성한다."""
    row: dict[str, Any] = {
        "A_거래처키": _safe_value(a_row.get("거래처키", "")),
        "A_거래처명": _safe_value(a_row.get("거래처명", "")),
        "A_거래처주소": _safe_value(a_row.get("거래처주소", "")),
        "A_전화번호": _safe_value(a_row.get("전화번호", "")),
        "매칭후보수": len(candidates),
        "최고점수": candidates[0].score.total_score if candidates else "",
        "매칭결과": "",
    }

    for i in range(1, config.MAX_CANDIDATES + 1):
        prefix = f"후보{i}"
        if i <= len(candidates):
            cand = candidates[i - 1]
            row[f"{prefix}_B거래처키"] = _safe_value(cand.b_row.get("거래처키", ""))
            row[f"{prefix}_B거래처명"] = _safe_value(cand.b_row.get("거래처명", ""))
            row[f"{prefix}_B거래처주소"] = _safe_value(cand.b_row.get("거래처주소", ""))
            row[f"{prefix}_B전화번호"] = _safe_value(cand.b_row.get("전화번호", ""))
            row[f"{prefix}_총점"] = cand.score.total_score
        else:
            row[f"{prefix}_B거래처키"] = ""
            row[f"{prefix}_B거래처명"] = ""
            row[f"{prefix}_B거래처주소"] = ""
            row[f"{prefix}_B전화번호"] = ""
            row[f"{prefix}_총점"] = ""

    return row


def get_output_columns() -> list[str]:
    """결과 파일 컬럼 순서."""
    cols = [
        "A_거래처키", "A_거래처명", "A_거래처주소", "A_전화번호",
        "매칭후보수", "최고점수", "매칭결과",
    ]
    for i in range(1, config.MAX_CANDIDATES + 1):
        cols.extend([
            f"후보{i}_B거래처키",
            f"후보{i}_B거래처명",
            f"후보{i}_B거래처주소",
            f"후보{i}_B전화번호",
            f"후보{i}_총점",
        ])
    return cols


def build_result_dataframe(df_a: pd.DataFrame, df_b: pd.DataFrame) -> pd.DataFrame:
    """A/B 데이터로 매칭 결과 DataFrame을 생성한다."""
    print(f"[INFO] A rows: {len(df_a)}, B rows: {len(df_b)}")
    print("[INFO] Normalizing B...")
    b_index = build_b_index(df_b)

    rows: list[dict[str, Any]] = []
    total = len(df_a)
    for i, (_, a_row) in enumerate(df_a.iterrows(), start=1):
        if i % 100 == 0 or i == total:
            print(f"[INFO] Matching row {i}/{total}")
        candidates = find_candidates_for_row(a_row, b_index)
        rows.append(_build_output_row(a_row, candidates))

    return pd.DataFrame(rows, columns=get_output_columns())


def _autosize_columns(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)


def _set_workbook_auto_calculation(workbook) -> None:
    """엑셀에서 파일을 열 때 수식이 다시 계산되도록 설정한다."""
    calculation = getattr(workbook, "calculation", None)
    if calculation is None:
        return
    calculation.calcMode = "auto"
    calculation.fullCalcOnLoad = True
    calculation.forceFullCalc = True


def _style_header(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill


def _write_criteria_sheet(writer) -> None:
    """매칭결과 구간 기준 sheet를 작성한다."""
    wb = writer.book
    ws = wb.create_sheet("기준")
    ws.append(["매칭결과", "시작점수", "종료점수"])
    for label, start, end in config.MATCH_RESULT_RANGES:
        ws.append([label, start, end])

    _style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row_idx in range(2, len(config.MATCH_RESULT_RANGES) + 2):
        ws[f"B{row_idx}"].number_format = "0"
        ws[f"C{row_idx}"].number_format = "0"
    _autosize_columns(ws)


def _matching_result_formula(score_cell: str) -> str:
    """기준 sheet의 구간표를 참조하는 매칭결과 수식을 만든다."""
    if not config.MATCH_RESULT_RANGES:
        return f'=IF({score_cell}="","미매칭","미매칭")'

    fallback = f"'기준'!$A${len(config.MATCH_RESULT_RANGES) + 1}"
    formula = fallback
    for idx in range(len(config.MATCH_RESULT_RANGES) - 1, -1, -1):
        criteria_row = idx + 2
        formula = (
            f"IF(AND({score_cell}>='기준'!$B${criteria_row},"
            f"{score_cell}<='기준'!$C${criteria_row}),"
            f"'기준'!$A${criteria_row},{formula})"
        )
    return f"=IF({score_cell}=\"\",{fallback},{formula})"


def _apply_matching_result_formulas(ws, df: pd.DataFrame) -> None:
    """매칭결과 컬럼에 엑셀 수식을 입력한다."""
    if "최고점수" not in df.columns or "매칭결과" not in df.columns:
        return

    score_col = df.columns.get_loc("최고점수") + 1
    result_col = df.columns.get_loc("매칭결과") + 1
    score_letter = get_column_letter(score_col)
    result_letter = get_column_letter(result_col)

    for row_idx in range(2, len(df) + 2):
        ws[f"{result_letter}{row_idx}"] = _matching_result_formula(
            f"${score_letter}{row_idx}"
        )


def _write_summary_sheet(writer, df: pd.DataFrame) -> None:
    """매칭결과별 건수와 비중 요약 sheet를 작성한다."""
    if "매칭결과" not in df.columns:
        return

    result_col = df.columns.get_loc("매칭결과") + 1
    result_letter = get_column_letter(result_col)
    ws = writer.book.create_sheet("요약")

    ws["A1"] = "매칭결과 요약"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["매칭결과", "건수", "비중"])

    start_row = 4
    for idx, _range in enumerate(config.MATCH_RESULT_RANGES):
        row_idx = start_row + idx
        criteria_row = idx + 2
        ws[f"A{row_idx}"] = f"='기준'!A{criteria_row}"
        ws[f"B{row_idx}"] = f'=COUNTIF(\'매칭결과\'!${result_letter}:${result_letter},A{row_idx})'

    total_row = start_row + len(config.MATCH_RESULT_RANGES)
    for row_idx in range(start_row, total_row):
        ws[f"C{row_idx}"] = f"=IF($B${total_row}=0,0,B{row_idx}/$B${total_row})"
        ws[f"C{row_idx}"].number_format = "0.0%"

    ws[f"A{total_row}"] = "합계"
    ws[f"B{total_row}"] = f"=SUM(B{start_row}:B{total_row - 1})"
    ws[f"C{total_row}"] = f"=IF(B{total_row}=0,0,1)"
    ws[f"C{total_row}"].number_format = "0.0%"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for cell in ws[total_row]:
        cell.font = Font(bold=True)
    _autosize_columns(ws)


def save_result_excel(df: pd.DataFrame, path: Path) -> None:
    """결과 엑셀을 스타일과 함께 저장한다."""
    _ensure_dir(path.parent)
    numeric_cols = {"매칭후보수", "최고점수"} | {
        f"후보{i}_총점" for i in range(1, config.MAX_CANDIDATES + 1)
    }

    try:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="매칭결과")
            ws = writer.sheets["매칭결과"]

            _style_header(ws)

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for col_name in numeric_cols:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1
                    letter = get_column_letter(col_idx)
                    for row_idx in range(2, len(df) + 2):
                        ws[f"{letter}{row_idx}"].number_format = "0"

            if "최고점수" in df.columns and "매칭결과" in df.columns:
                _write_criteria_sheet(writer)
                _apply_matching_result_formulas(ws, df)
                _write_summary_sheet(writer, df)
            _set_workbook_auto_calculation(writer.book)
            _autosize_columns(ws)
    except PermissionError:
        raise PermissionError(
            f"결과 파일을 저장할 수 없습니다: {path}\n"
            "파일이 다른 프로그램(엑셀 등)에서 열려 있을 수 있습니다. "
            "파일을 닫은 후 다시 실행해 주세요."
        )
