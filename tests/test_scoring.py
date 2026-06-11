"""점수 계산 및 결과 출력 테스트."""

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

import config
from matcher.address_normalizer import normalize_address
from matcher.candidate import _collect_indices, build_b_index, find_candidates_for_row
from matcher.excel_io import build_result_dataframe, get_output_columns, save_result_excel
from matcher.name_normalizer import normalize_name
from matcher.phone_normalizer import normalize_phone
from matcher.scoring import calc_address_score, calc_phone_bonus, calc_total_score


def _make_name(text: str):
    return normalize_name(text)


def _make_addr(text: str):
    return normalize_address(text)


def test_same_name_high_score():
    name = "가톨릭대학교 서울성모병원"
    score = calc_total_score(
        _make_name(name),
        _make_name(name),
        _make_addr("서울특별시 서초구 반포대로 222"),
        _make_addr("서울특별시 서초구 반포대로 222"),
        "0212345678",
        "0212345678",
    )
    assert score.name_score >= 42


def test_same_address_high_score():
    addr = "서울특별시 강남구 테헤란로 152"
    score = calc_total_score(
        _make_name("서울병원"),
        _make_name("서울의원"),
        _make_addr(addr),
        _make_addr(addr),
        "",
        "",
    )
    assert score.address_score >= 45


def test_same_phone_bonus():
    assert calc_phone_bonus("0212345678", "0212345678") == 20


def test_total_score_capped_at_100():
    score = calc_total_score(
        _make_name("가톨릭대학교 서울성모병원"),
        _make_name("가톨릭대학교 서울성모병원"),
        _make_addr("서울특별시 서초구 반포대로 222"),
        _make_addr("서울특별시 서초구 반포대로 222"),
        "0212345678",
        "0212345678",
    )
    assert score.total_score <= 100


def test_low_score_candidates_excluded_when_threshold_is_high(monkeypatch):
    monkeypatch.setattr(config, "MATCH_THRESHOLD", 80)
    df_a = pd.DataFrame([{
        "거래처키": "A001",
        "거래처명": "완전다른병원AAA",
        "거래처주소": "서울특별시 강남구 테헤란로 1",
        "전화번호": "0211111111",
    }])
    df_b = pd.DataFrame([{
        "거래처키": "B001",
        "거래처명": "전혀다른의원ZZZ",
        "거래처주소": "부산광역시 해운대구 해운대로 999",
        "전화번호": "0519999999",
    }])

    b_index = build_b_index(df_b)
    candidates = find_candidates_for_row(df_a.iloc[0], b_index, fallback_all=True)
    assert len(candidates) == 0


def test_candidate_index_uses_region_key_not_sigungu_only():
    df_a = pd.DataFrame([{
        "거래처키": "A001",
        "거래처명": "옥천병원",
        "거래처주소": "부산광역시 서구 옥천로 10",
        "전화번호": "051-111-2222",
    }])
    df_b = pd.DataFrame([
        {
            "거래처키": "B001",
            "거래처명": "광주옥천병원",
            "거래처주소": "광주광역시 서구 옥천로 10",
            "전화번호": "062-111-2222",
        },
        {
            "거래처키": "B002",
            "거래처명": "옥천병원",
            "거래처주소": "부산광역시 서구 옥천로 10",
            "전화번호": "051-111-2222",
        },
    ])

    b_index = build_b_index(df_b)
    assert set(b_index.by_region) == {"광주|서구", "부산|서구"}

    candidates = find_candidates_for_row(df_a.iloc[0], b_index)
    assert [c.b_row["거래처키"] for c in candidates] == ["B002"]


def test_fallback_candidate_collection_has_no_hard_limit():
    row_count = 1005
    df_b = pd.DataFrame(
        [
            {
                "거래처키": f"B{i:04d}",
                "거래처명": f"후보병원{i}",
                "거래처주소": "",
                "전화번호": "",
            }
            for i in range(row_count)
        ]
    )
    b_index = build_b_index(df_b)

    positions = _collect_indices(
        b_index,
        normalize_name(""),
        normalize_address(""),
        fallback_all=True,
    )

    assert len(positions) == row_count


@pytest.fixture
def sample_dataframes():
    df_a = pd.DataFrame([
        {
            "거래처키": "A001",
            "거래처명": "서울성모병원",
            "거래처주소": "서울특별시 서초구 반포대로 222",
            "전화번호": "02-1234-5678",
        },
        {
            "거래처키": "A002",
            "거래처명": "완전다른병원",
            "거래처주소": "제주특별자치도 제주시 오라동 1",
            "전화번호": "",
        },
    ])
    df_b = pd.DataFrame([
        {
            "거래처키": "B001",
            "거래처명": "서울성모병원",
            "거래처주소": "서울특별시 서초구 반포대로 222",
            "전화번호": "02-1234-5678",
        },
        {
            "거래처키": "B002",
            "거래처명": "서울성모병원",
            "거래처주소": "서울특별시 서초구 반포대로 220",
            "전화번호": "02-1234-5678",
        },
    ])
    return df_a, df_b


def test_result_excludes_internal_columns(sample_dataframes, tmp_path):
    df_a, df_b = sample_dataframes
    result = build_result_dataframe(df_a, df_b)
    cols = set(result.columns)

    forbidden_prefixes = [
        "A_정규화", "B_정규화", "B원본행번호",
        "거래처명점수", "주소점수", "전화번호가점",
        "판정", "검토사유",
    ]
    for col in cols:
        for prefix in forbidden_prefixes:
            assert prefix not in col
        assert not col.endswith("_표시")

    assert set(get_output_columns()) == cols


def test_result_contains_matching_result_column(sample_dataframes):
    df_a, df_b = sample_dataframes
    result = build_result_dataframe(df_a, df_b)
    cols = list(result.columns)
    assert "매칭결과" in cols
    assert cols.index("매칭결과") == cols.index("최고점수") + 1


def test_result_saved_without_extra_columns(sample_dataframes, tmp_path):
    df_a, df_b = sample_dataframes
    result = build_result_dataframe(df_a, df_b)
    out_path = tmp_path / "C.xlsx"
    save_result_excel(result, out_path)
    loaded = pd.read_excel(out_path)
    assert "A_정규화거래처명" not in loaded.columns
    assert "판정" not in loaded.columns


def test_result_excel_has_summary_and_criteria_sheets(sample_dataframes, tmp_path):
    df_a, df_b = sample_dataframes
    result = build_result_dataframe(df_a, df_b)
    out_path = tmp_path / "C.xlsx"
    save_result_excel(result, out_path)

    workbook = load_workbook(out_path, data_only=False)
    assert {"매칭결과", "요약", "기준"}.issubset(set(workbook.sheetnames))

    result_ws = workbook["매칭결과"]
    headers = [cell.value for cell in result_ws[1]]
    score_col = headers.index("최고점수") + 1
    result_col = headers.index("매칭결과") + 1
    assert result_col == score_col + 1
    assert str(result_ws.cell(row=2, column=result_col).value).startswith("=IF(")

    criteria_ws = workbook["기준"]
    assert criteria_ws["A1"].value == "매칭결과"
    assert criteria_ws["A2"].value == "확정"
    assert criteria_ws["B2"].value == 80
    assert criteria_ws["C2"].value == 100

    summary_ws = workbook["요약"]
    assert summary_ws["A1"].value == "매칭결과 요약"
    assert "COUNTIF" in str(summary_ws["B4"].value)
    assert str(summary_ws["C4"].value).startswith("=IF(")
